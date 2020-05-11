#!/usr/bin/env python
# coding: utf-8

# In[9]:
# VaI with SMC implementation

import numpy as np
import tensorflow as tf
import math
import h5py
import matplotlib.pyplot as plt
from tensorflow.python.framework import ops
from tensorflow.contrib import rnn
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import scipy.stats
from math import sqrt
import random
from sklearn.metrics import mean_squared_error

conf = tf.ConfigProto(allow_soft_placement=True)
conf.gpu_options.per_process_gpu_memory_fraction = 0.95
conf.gpu_options.allow_growth = True
os.environ['CUDA_VISIBLE_DEVICES'] = '0'

class VAE_Network(object):
    def __init__(self):
        self.hidden = 3
        self.X = tf.placeholder(tf.float32, [None, 3],name='Placeholder_X')
        self.Y = tf.placeholder(tf.float32, [None, 1],name='Placeholder_Y')
        self.Xt = tf.placeholder(tf.float32, [None, 3],name='Placeholder_Xt')
        self.X1 = tf.placeholder(tf.float32, [None, 3],name='Placeholder_X1')
        self.Y1 = tf.placeholder(tf.float32, [None, 1],name='Placeholder_Y1')
        self.Xt1 = tf.placeholder(tf.float32, [None, 3],name='Placeholder_Xt1')
        
        with tf.variable_scope('Encoder_Transition'):
            q_mu, q_sigma = self.inference_network(self.X, self.Y)
            eps = tf.random_normal(shape=tf.shape(q_mu))
            self.Z = tf.multiply(eps, q_sigma) + q_mu
            q_z = tf.distributions.Normal(loc=q_mu, scale=q_sigma)
        with tf.variable_scope('Decoder_Transition'):
            p_x_given_z_mu, p_xsig = self.generative_network(self.X, self.Z)
            p_x_given_z_dist = tf.distributions.Normal(loc=p_x_given_z_mu, scale=p_xsig)
            self.Yout = p_x_given_z_mu
        with tf.variable_scope('Prior_Transition'):
            p_z = tf.distributions.Normal(loc=np.zeros(self.hidden, dtype=np.float32),scale=np.ones(self.hidden, dtype=np.float32))
        with tf.variable_scope('Decoder_Transition',reuse=True):
            self.Zt = tf.random_normal(shape=tf.shape(q_mu))
            self.posterior, self.dummy = self.generative_network(self.Xt, self.Zt)
        
        with tf.variable_scope('Encoder1_Sensor'):
            q_mu1, q_sigma1 = self.inference_network1(self.X1, self.Y1)
            eps1 = tf.random_normal(shape=tf.shape(q_mu1))
            self.Z1 = tf.multiply(eps1, q_sigma1) + q_mu1
            q_z1 = tf.distributions.Normal(loc=q_mu1, scale=q_sigma1)
        with tf.variable_scope('Decoder1_Sensor'):
            p_x_given_z_mu1, p_ssig = self.generative_network1(self.X1, self.Z1)
            p_x_given_z_dist1 = tf.distributions.Normal(loc=p_x_given_z_mu1, scale=p_ssig)
            self.Yout1 = p_x_given_z_mu1
        with tf.variable_scope('Prior1_Sensor'):
            p_z1 = tf.distributions.Normal(loc=np.zeros(self.hidden, dtype=np.float32),scale=np.ones(self.hidden, dtype=np.float32))
        with tf.variable_scope('Decoder1_Sensor',reuse=True):
            self.Zt1 = tf.random_normal(shape=tf.shape(q_mu1))
            self.posterior1, self.dummy1 = self.generative_network1(self.Xt1, self.Zt1)

        self.expected_log_likelihood = tf.reduce_sum(p_x_given_z_dist.log_prob(self.Y))
        self.kl = tf.reduce_sum(tf.distributions.kl_divergence(q_z,p_z))
        self.elbo = tf.reduce_mean(self.expected_log_likelihood -self.kl)
        self.lr = 3e-5
        self.global_step = tf.Variable(0, trainable = False)
        self.decay = tf.train.exponential_decay(self.lr, self.global_step, 100, 0.99, staircase = True)
        self.optimizer = tf.train.AdamOptimizer(self.decay, name='Adam_Transition').minimize(-self.elbo)

        self.expected_log_likelihood1 = tf.reduce_sum(p_x_given_z_dist1.log_prob(self.Y1))
        self.kl1 = tf.reduce_sum(tf.distributions.kl_divergence(q_z1,p_z1))
        self.elbo1 = tf.reduce_mean(self.expected_log_likelihood1 -self.kl1)
        self.lr1 = 1e-5
        self.global_step1 = tf.Variable(0, trainable = False)
        self.decay1 = tf.train.exponential_decay(self.lr1, self.global_step1, 100, 0.99, staircase = True)
        self.optimizer1 = tf.train.AdamOptimizer(self.decay1, name='Adam_Sensor').minimize(-self.elbo1)
        

    def inference_network(self, X, Y):
        self.Encoder_Input = tf.concat([X, Y], 1)
        self.HiddenE1 = tf.contrib.layers.fully_connected(self.Encoder_Input, 250, activation_fn = tf.nn.leaky_relu, 
                                                                 weights_initializer = tf.contrib.layers.xavier_initializer())
        self.HiddenE2 = tf.contrib.layers.fully_connected(self.HiddenE1, 60, activation_fn = tf.nn.leaky_relu, 
                                                                 weights_initializer = tf.contrib.layers.xavier_initializer())
        self.EncoderOut = tf.contrib.layers.fully_connected(self.HiddenE2, 2*self.hidden, activation_fn = None,
                                                                 weights_initializer = tf.contrib.layers.xavier_initializer())
        
        mui, sigmai = tf.split(self.EncoderOut, num_or_size_splits=2, axis=1)
        return mui, sigmai
    
    def generative_network(self, X, Z):
        self.Decoder_Input = tf.concat([X, Z], 1)
        self.HiddenD1 = tf.contrib.layers.fully_connected(self.Decoder_Input, 250, activation_fn = tf.nn.leaky_relu, 
                                                                 weights_initializer = tf.contrib.layers.xavier_initializer())
        self.HiddenD2 = tf.contrib.layers.fully_connected(self.HiddenD1, 60, activation_fn = tf.nn.leaky_relu, 
                                                                 weights_initializer = tf.contrib.layers.xavier_initializer())
        self.DecoderOut = tf.contrib.layers.fully_connected(self.HiddenD2, 2, activation_fn =  None, 
                                                                 weights_initializer = tf.contrib.layers.xavier_initializer())
        mug, sigmag = tf.split(self.DecoderOut, num_or_size_splits=2, axis=1)
        return mug, 1+sigmag**2
        
    def inference_network1(self, X1, Y1):
        self.Encoder_Input1 = tf.concat([X1, Y1], 1)
        self.HiddenE11 = tf.contrib.layers.fully_connected(self.Encoder_Input1, 250, activation_fn = tf.nn.leaky_relu, 
                                                                 weights_initializer = tf.contrib.layers.xavier_initializer())
        self.HiddenE21 = tf.contrib.layers.fully_connected(self.HiddenE11, 60, activation_fn = tf.nn.leaky_relu, 
                                                                 weights_initializer = tf.contrib.layers.xavier_initializer())
        self.EncoderOut1 = tf.contrib.layers.fully_connected(self.HiddenE21, 2*self.hidden, activation_fn = None,
                                                                 weights_initializer = tf.contrib.layers.xavier_initializer())
        
        mui1, sigmai1 = tf.split(self.EncoderOut1, num_or_size_splits=2, axis=1)
        return mui1, sigmai1
    
    def generative_network1(self, X1, Z1):
        self.Decoder_Input1 = tf.concat([X1, Z1], 1)
        self.HiddenD11 = tf.contrib.layers.fully_connected(self.Decoder_Input1, 250, activation_fn = tf.nn.leaky_relu, 
                                                                 weights_initializer = tf.contrib.layers.xavier_initializer())
        self.HiddenD21 = tf.contrib.layers.fully_connected(self.HiddenD11, 60, activation_fn = tf.nn.leaky_relu, 
                                                                 weights_initializer = tf.contrib.layers.xavier_initializer())
        self.DecoderOut1 = tf.contrib.layers.fully_connected(self.HiddenD21, 2, activation_fn =  None, 
                                                                 weights_initializer = tf.contrib.layers.xavier_initializer())
        mug1, sigmag1 = tf.split(self.DecoderOut1, num_or_size_splits=2, axis=1)
        return mug1, 1+sigmag1**2
    
def resample(weights):
    n = len(weights)
    indices = []
    C = [0.] + [sum(weights[:i+1]) for i in range(n)]
    u0 = random.random()
    j = 0
    for u in [(u0+i)/n for i in range(n)]:
        while u > C[j]:
            j+=1
        indices.append(j-1)
    return indices

def load_data():
    wb_Tr = Workbook()
    wb_Tr = load_workbook(filename = 'HBN_VI4.xlsx', data_only = True)
    ws_Tr = wb_Tr.active

    Train_num = 6830
    Test_num = 4400

    UX_train1 = np.zeros((Train_num,4))
    XY_train1 = np.zeros((Train_num,4))
    U_test1 = np.zeros((Test_num,1))
    X_test1 = np.zeros((Test_num,1))
    Y_test1 = np.zeros((Test_num,1))

    for i in range(Train_num):
        for j in range(4):
            UX_train1[i][j] = ws_Tr.cell(row=i+1, column=j+1).value
            XY_train1[i][j] = ws_Tr.cell(row=i+1, column=j+5).value
    for i in range(Test_num):
        U_test1[i][0] = ws_Tr.cell(row=i+1, column=9).value
        X_test1[i][0] = ws_Tr.cell(row=i+1, column=12).value
        Y_test1[i][0] = ws_Tr.cell(row=i+1, column=11).value

    UX_train1[:,0] = UX_train1[:,0]*100
    UX_train1[:,1] = UX_train1[:,1]*100
    UX_train1[:,2] = UX_train1[:,2]-250
    UX_train1[:,3] = UX_train1[:,3]-250
    XY_train1[:,0] = XY_train1[:,0]-250
    XY_train1[:,1] = XY_train1[:,1]-250
    XY_train1[:,2] = XY_train1[:,2]*100000 - 9963
    XY_train1[:,3] = XY_train1[:,3]*100000 - 9963
    Y_test1 = Y_test1 * 100000 - 9963
    U_test1 = U_test1 * 100
    X_test1 = X_test1 - 250
    return UX_train1, XY_train1, U_test1, X_test1, Y_test1


def VI_train(UX_train1, XY_train1, conf):
    minibatch_size = 256
    tf.reset_default_graph()
    Dymodel = VAE_Network()
    
    Z = UX_train1
    Z = np.random.permutation(Z)
    X = Z[:,0:3]
    Y = Z[:,3]
    Y = np.reshape(Y, [-1,1])
    with tf.Session(config=conf) as sess:
        sess.run(tf.global_variables_initializer())
        sess.run(tf.local_variables_initializer())
        for epoch in range(2501):
            Lower_Bound = 0
            A = 0
            B = 0
            for i in range((int)((np.shape(X)[0])/minibatch_size)):
                X_batch = X[i*minibatch_size:(i+1)*minibatch_size, :]
                Y_batch = Y[i*minibatch_size:(i+1)*minibatch_size, :]
                _,ci, a, b = sess.run([Dymodel.optimizer, Dymodel.elbo, Dymodel.kl, Dymodel.expected_log_likelihood], feed_dict={Dymodel.X:X_batch, Dymodel.Y:Y_batch})
                Lower_Bound = Lower_Bound + ci
                A = A + a
                B = B + b
            Lower_Bound = Lower_Bound / (int)(np.shape(X)[0]/minibatch_size)
            A = A / (int)(np.shape(X)[0]/minibatch_size)
            B = B / (int)(np.shape(X)[0]/minibatch_size)
            if epoch % 100 == 0 or epoch <= 3:
                print("Training","Epoch:",epoch,"Lower Bound:",Lower_Bound, "KL:",A,"ELL:",B)
        
    
        Z = XY_train1
        Z = np.random.permutation(Z)
        X = Z[:,0:3]
        Y = Z[:,3]
        Y = np.reshape(Y, [-1,1])
        for epoch in range(4001):
            Lower_Bound = 0
            A = 0
            B = 0
            for i in range((int)((np.shape(X)[0])/minibatch_size)):
                X_batch = X[i*minibatch_size:(i+1)*minibatch_size, :]
                Y_batch = Y[i*minibatch_size:(i+1)*minibatch_size, :]
                _,ci, a, b = sess.run([Dymodel.optimizer1, Dymodel.elbo1 , Dymodel.kl1, Dymodel.expected_log_likelihood1], feed_dict={Dymodel.X1:X_batch, Dymodel.Y1:Y_batch})
                Lower_Bound = Lower_Bound + ci
                A = A + a
                B = B + b
            Lower_Bound = Lower_Bound / (int)(np.shape(X)[0]/minibatch_size)
            A = A / (int)(np.shape(X)[0]/minibatch_size)
            B = B / (int)(np.shape(X)[0]/minibatch_size)
            #if A <= 1:
            #    break
            if epoch % 100 == 0 or epoch <= 3:
                print("Training","Epoch:",epoch,"Lower Bound:",Lower_Bound, "KL:",A,"ELL:",B)
        ckpt_dir = './ckpt/'
        if not os.path.exists(ckpt_dir):
            os.makedirs(ckpt_dir)
        ckpt_path = ckpt_dir + 'VI_train_weight' + '.ckpt'
        saver = tf.train.Saver()
        saver.save(sess, ckpt_path)       
        
def demo_test(U_test1, X_test1, Y_test1, conf):
    tf.reset_default_graph()
    with tf.Session(config=conf) as sess:
        fd = 50
        dn = 50
        Dymodel = VAE_Network()
        ckpt_dir = './ckpt/'
        ckpt_path_location = ckpt_dir + 'VI_train_weight' + '.ckpt'
        saver = tf.train.Saver()
        saver.restore(sess, ckpt_path_location)
        X = np.zeros((fd,1))
        Xpred = np.zeros((fd,1))
        Xprev = np.zeros((fd,1))
        Weight = np.zeros((fd,1))
        X_predict = np.zeros((4400,1))
        X_predict[0,0] = 84
        for particle in range(fd):
            Xpred[particle][0] = 84
            X[particle][0] = 84 + 1*(random.random() - 0.5)
        for epoch in range(1,4400):
            if epoch % 100 == 1:
                print(epoch)
            Xprev = X
            Dyinput = np.concatenate((np.ones((fd,1))*U_test1[epoch,:], np.ones((fd,1))*U_test1[epoch-1,:], Xprev),1)
            Dydummy = np.zeros((dn,3))
            Dyoutput = np.zeros((dn,1))
            Xpredd = sess.run([Dymodel.posterior], feed_dict={Dymodel.X: Dyinput, Dymodel.Y: Dyoutput, Dymodel.Xt: Dyinput})
            Xpredd = np.array(Xpredd)
            Xpred = np.reshape(Xpredd[:,:,0],[-1,1])

            Weight = np.zeros((fd,1))
            for j in range(1):
                Seinput = np.concatenate((Xpred, Xprev, np.ones((fd,1))*Y_test1[epoch-1,0]),1)
                Seoutput = np.zeros((fd,1))
                Ypredd, Ys = sess.run([Dymodel.posterior1, Dymodel.dummy1], feed_dict={Dymodel.X1: Seinput, Dymodel.Y1: Seoutput, Dymodel.Xt1: Seinput})
                Ypredd = np.reshape(np.array(Ypredd),[-1])
                Ys=  np.reshape(np.array(Ys),[-1])
                for particle in range(fd):
                    Weight[particle,0] = Weight[particle,0] + scipy.stats.norm(Ypredd[particle], Ys[particle]).pdf(Y_test1[epoch,0])
            Weight = Weight/sum(Weight)
            X = Xpred[resample(Weight)]
            X_predict[epoch,0] = sum(X)/fd
    return X_predict

    
if __name__ == '__main__':
    UX_train1, XY_train1, U_test1, X_test1, Y_test1 = load_data()
    VI_train(UX_train1, XY_train1, conf)
    X_predict = demo_test(U_test1, X_test1, Y_test1, conf)


# In[ ]:





# In[ ]:





# In[ ]:




